import os

import openpyxl
import csv
import xml.etree.ElementTree as ET


class VisionBox:
    MACHINE_SPEED = None
    C1_GRIPPERS = None
    C2_CLAMPS = None
    VLK_CAMPANE = None

    def __init__(self, name):
        # Inițializează liste pentru interfețele de rețea
        self.name = name
        self.eth0 = []
        self.eth1 = []
        self.eth2 = []
        self.eth3 = []
        self.eth4 = []
        self.stations = {}
        self.number_of_stations = 0

# adaugam statiitle nume, ordinea pentru station name, ordinea gloabala,pozitia in masina,nr spin

    def set_machine_characteristics(cls, **kwargs):
        for key, value in kwargs.items():
            if value is not None and hasattr(cls, key):
                setattr(cls, key, value)



    def add_station(self, station_name, station_order, stglobalorder, tipst, spinnr=''):
        self.stations[station_name] = [station_order, stglobalorder, tipst, spinnr]
        self.number_of_stations += 1

    def add_ip(self, ip):  # assignam adresele ip
        for name, value in ip.items():
            if name == 'ETH0[IP]':
                for ipadd in value:
                    self.eth0.append(ipadd)
            if name == 'ETH1[IP]':
                for ipadd in value:
                    self.eth1.append(ipadd)
            if name == 'ETH2[IP]':
                for ipadd in value:
                    self.eth2.append(ipadd)
            if name == 'ETH3[IP]':
                for ipadd in value:
                    self.eth3.append(ipadd)
            if name == 'ETH4[IP]':
                for ipadd in value:
                    self.eth4.append(ipadd)

    def showip(self):  # afisam ip-urile
        print('ETHO: {}'.format(self.eth0))
        print('ETH1: {}'.format(self.eth1))
        print('ETH2: {}'.format(self.eth2))
        print('ETH3: {}'.format(self.eth3))
        print('ETH4: {}'.format(self.eth4))

    def add_tags_from_list(self, datatags, path, station_name):
        xml_file_path = os.path.join(path, f"{station_name}_InstallParam.xml")

        # Verifică dacă fișierul există
        if os.path.exists(xml_file_path):
            tree = ET.parse(xml_file_path)
            root = tree.getroot()
        else:
            # Creează un arbore nou dacă fișierul nu există
            root = ET.Element("Root")
            tree = ET.ElementTree(root)

        for items in datatags:
            tag_tree = create_tag_list(items['VariableName'], items['PlcName'].format(self.stations[station_name][1]),
                                       items['type'])
            root.append(tag_tree)

        ET.indent(root, space='  ', level=0)
        tree.write(xml_file_path, encoding="utf-8", xml_declaration=False, method="xml", short_empty_elements=False)

    def writetags(self, datatags, datatags1, path):
        if not os.path.exists(path):
            os.mkdir(path)

        for numest, st in self.stations.items():
            base_tree = ET.ElementTree(ET.Element("Root"))
            for items in datatags:
                tag_tree = create_tag_list(items['VariableName'], items['PlcName'].format(st[1]), items['type'])
                base_tree.getroot().append(tag_tree)

            # Adauga tag-urile pentru data1
            for items1 in datatags1:
                tag_tree1 = create_tag_list(items1['VariableName'], items1['PlcName'].format(st[3]), items1['type'])
                base_tree.getroot().append(tag_tree1)

            ET.indent(base_tree, space='  ', level=0)
            base_tree.write(os.path.join(path, f"{numest}_InstallParam.xml"), encoding="utf-8", xml_declaration=False,
                            method="xml", short_empty_elements=False)
    def afiseazaconst(self):
        print('{}{}{}{}'.format(self.VLK_CAMPANE, self.C1_GRIPPERS, self.MACHINE_SPEED, self.C2_CLAMPS))

def read_from_excel(excelfile, startrangeip, endrangeip, numarul_coloanei):  # citim ip-uriel din excel file
    source_excel_file_path = excelfile
    work_book = openpyxl.load_workbook(source_excel_file_path)
    work_sheet = work_book['Sheet1']
    list_ip_addr = []
    for actual_row_number in range(startrangeip, endrangeip + 1):
        value_cell = work_sheet.cell(row=actual_row_number, column=numarul_coloanei).value
        if value_cell is not None:
            list_ip_addr.append(value_cell)
    work_book.close()
    return list_ip_addr


def write_to_file(vision_box, destination_path_file):
    if os.path.exists(destination_path_file):
        pass
    else:
        os.mkdir(destination_path_file)
    with open(destination_path_file+'/config.txt', 'w') as file:
        if not vision_box.eth0[0]:
            pass
        else:
            file.write('#---------ETH0------------\n')
            file.write('{}\n'.format(vision_box.eth0[0]))
        if len(vision_box.eth0)<2:
            pass
        else:
            file.write('# -------ETH0 Alias--------\n')
            for count in range(1, len(vision_box.eth0)):
                file.write('{}\n'.format(vision_box.eth0[count]))
        if not vision_box.eth1[0]:
            pass
        else:
            file.write('#---------ETH1------------\n')
            file.write('{}\n'.format(vision_box.eth1[0]))
        if not vision_box.eth2:
            pass
        else:
            file.write('#---------ETH2------------\n')
            file.write('{}\n'.format(vision_box.eth2[0]))
        if not vision_box.eth3:
            pass
        else:
            file.write('#---------ETH3------------\n')
            file.write('{}\n'.format(vision_box.eth3[0]))
        if not vision_box.eth4:
            pass
        else:
            file.write('#---------ETH4------------\n')
            file.write('{}\n'.format(vision_box.eth4[0]))


def cautaindict(sursa, val):  # o folosesc pentru a gasi statiunile de pe carousel 1 si a le pune in ordine
    sursad = sursa
    value_to_find = val[:3]
    found_in_key = None
    # Parcurge fiecare cheie și lista asociată pentru a găsi valoarea
    for key, value_list in sursad.items():
        if value_to_find.upper() in value_list:
            found_in_key = key
            break
    if found_in_key is not None:
        return found_in_key
    else:
        print(f"Valoarea '{value_to_find}' nu a fost găsită în nicio listă asociată cu cheie.")


def create_tag_list(root_element, plcname, datatype):  # crearea tagului  este chemata mai sus in metoda writetag
    root = ET.Element(root_element)
    plc_var_name = ET.SubElement(root, "PLCVarName", AuditItemType="-1",
                                 Description="Name of variable defined in the PLC")
    plc_var_name.text = plcname
    db_number = ET.SubElement(root, "DBNumber", AuditItemType="-1", Description="Memory block")
    db_number.text = "9"
    byte_offset = ET.SubElement(root, "ByteOffset", AuditItemType="-1", Description="Byte Offset")
    byte_offset.text = "120"
    bit_offset = ET.SubElement(root, "BitOffset", AuditItemType="-1", Description="Bit Offset")
    bit_offset.text = "0"
    data_type = ET.SubElement(root, "Type", AuditItemType="-1", Description="Data type")
    data_type.text = datatype
    return root

def data_excel_validation(value_to_validate):
    if value_to_validate:
        return value_to_validate[0]
    else:
        return None


def read_csv_file(file_path):
    tempdata = []
    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:  # Deschide fișierul CSV și citește conținutul
        csvreader = csv.DictReader(csvfile)
        for row in csvreader:  # Iterează prin rândurile din fișierul CSV
            tempdata.append(row)
    return tempdata


workingPath = os.getcwd()
pathStart = workingPath + '//Output'
if os.path.exists(pathStart):
    pass
else:
    os.mkdir(pathStart)
excel_file_path = 'ConfigFiles/ConfigExcel.xlsm'

number_of_vb = data_excel_validation(read_from_excel(excel_file_path, 22, 22, 6))
macchine_speed = data_excel_validation(read_from_excel(excel_file_path, 23, 23, 6))
c1_grippers = data_excel_validation(read_from_excel(excel_file_path, 24, 24, 6))
c2_clamps = data_excel_validation(read_from_excel(excel_file_path, 25, 25, 6))
leak_campane = data_excel_validation(read_from_excel(excel_file_path, 26, 26, 6))
plc_type = data_excel_validation(read_from_excel(excel_file_path, 27, 27, 6))

VisionBox.set_machine_characteristics(VisionBox, MACHINE_SPEED=macchine_speed, C1_GRIPPERS=c1_grippers, C2_CLAMPS=c2_clamps, VLK_CAMPANE=leak_campane )
columnarray = []
initcol = 2
for i in range(0, number_of_vb):
    columnarray.append(initcol)
    initcol += 2
column_number = 2
startrangeIp = 5
endrangeIp = 13
dictIp = {}
vbarr = []
for i in range(0, number_of_vb+1):
    creare_vb = VisionBox(str(i+1))
    vbarr.append(creare_vb)
# Citirea ip-urilor din excel file i fiind coloana, de fiecare data pleaca la urmatoarea si cheama metoda readips
k = 0
for i in columnarray:
    dictIp['ETH0[IP]'] = read_from_excel(excel_file_path, 5, 13, i)
    dictIp['ETH1[IP]'] = read_from_excel(excel_file_path, 14, 14, i)
    dictIp['ETH2[IP]'] = read_from_excel(excel_file_path, 15, 15, i)
    dictIp['ETH3[IP]'] = read_from_excel(excel_file_path, 16, 16, i)
    dictIp['ETH4[IP]'] = read_from_excel(excel_file_path, 17, 17, i)
    vbarr[k].add_ip(dictIp)
    k += 1
# Panel PC assign Ip separat dara in acelasi array
dictIp.clear()
dictIp['ETH0[IP]'] = read_from_excel(excel_file_path, 5, 13, 12)
dictIp['ETH1[IP]'] = read_from_excel(excel_file_path, 14, 14, 12)
vbarr[k].add_ip(dictIp)
vbarr[k].name = 'PanelPC'
###############################################################

for i in vbarr:
    i.showip()
#  Scriere ip in fisiere
for i in vbarr:
    if i.name == 'PanelPC':
        write_to_file(i, pathStart+'//'+i.name)
    else:
        write_to_file(i, pathStart + '//' + 'VB' + i.name)

# deschidem fisierul excel si citim statiunile si ordinea lor
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook['Sheet1']
statdict = {}
listorder = []
for row_number in range(23, 45 + 1):
    cell_value = sheet.cell(row=row_number, column=column_number).value
    if cell_value is not None:
        statdict[cell_value] = [sheet.cell(row=row_number, column=column_number-1).value, sheet.cell(row=row_number,
                                                        column=column_number+1).value]
        listorder.append(cell_value)
workbook.close()
#####################################################################
# deschidem fisierul si verificam fiecare statie daca se afla in lista statiilor de pe c1,c2,leak,single
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook['Sheet2']
pozDict = {}
temp_list = []
for row_number in range(3, 25 + 1):
    cell_value = sheet.cell(row=row_number, column=1).value
    if cell_value is not None:
        temp_list.append(cell_value)
pozDict['C1'] = temp_list
temp_list = []
for row_number in range(3, 25 + 1):
    cell_value = sheet.cell(row=row_number, column=2).value
    if cell_value is not None:
        temp_list.append(cell_value)
pozDict['C2'] = temp_list
temp_list = []
for row_number in range(3, 25 + 1):
    cell_value = sheet.cell(row=row_number, column=3).value
    if cell_value is not None:
        temp_list.append(cell_value)
pozDict['VLK'] = temp_list
temp_list = []
for row_number in range(3, 25 + 1):
    cell_value = sheet.cell(row=row_number, column=4).value
    if cell_value is not None:
        temp_list.append(cell_value)
pozDict['Single'] = temp_list
workbook.close()
################################################################
# cream dictioanrul gen statia : nrspin
c1 = {}
k = 1
for i in listorder:
    if i[:3] in pozDict['C1']:
        c1[i] = round(k/2+0.1)
        k += 1
#########################################
# adaugam stattile, ordinea,pozitia(single,c1,c2,leak)
for vb in vbarr:
    for stname, orderlist in statdict.items():
        if str(orderlist[1]) == str(vb.name):
            poz = cautaindict(pozDict, stname)
            if poz == 'C1':
                vb.add_station(stname, orderlist[0], listorder.index(stname)+1, poz, str(c1[stname]))
            else:
                vb.add_station(stname, orderlist[0], listorder.index(stname) + 1, poz)
###############################################################
file_path = 'ConfigFiles/InstalParamVariable.csv'
file_pathspins = 'ConfigFiles/SpinsVariable.csv'
data = []
data1 = []
data1 = read_csv_file(file_pathspins)
data = read_csv_file(file_path)


for visionb in vbarr:
    if visionb.name == 'PanelPC':
        visionb.writetags(data, data1, pathStart + '//' + visionb.name)
    else:
        visionb.writetags(data, data1, pathStart + '//' + 'VB' + visionb.name)

print(vbarr[0].stations)